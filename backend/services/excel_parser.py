"""
Excel workbook parser for vendor field mapping and charge mapping.

Workbook structure (one file, multiple sheets):
  • One tab per vendor_ref_id (e.g. "CLIM") — field mapping rules
  • A "Charge Map" tab — charge code → vendor-specific name lookup

Both lookups are keyed by vendor_reference_id from the invoice payload.
"""

import io

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _cell(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def parse_field_mapping(excel_bytes: bytes, vendor_ref_id: str) -> list[dict]:
    """
    Open the sheet named <vendor_ref_id> and return field mapping rules.

    Returns:
        [
          {
            "field_name":       "invoice_number",
            "payload_field":    "invoice_number",
            "ingestion_logic":  "Extraction Logic",
            "mapping_required": "No",
            "remarks":          "E.g.: Invoice #: 2390575",
            "value_present":    "Available in Payload"
          },
          ...
        ]
    """
    if not _HAS_OPENPYXL:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if name.strip().upper() == vendor_ref_id.strip().upper():
            sheet = wb[name]
            break

    if sheet is None:
        return []

    all_rows = list(sheet.iter_rows())

    # Find the header row by scanning for the "Column Name" keyword
    header_idx = None
    col_map: dict[str, int] = {}

    for i, row in enumerate(all_rows):
        vals = [_cell(c).lower() for c in row]
        if any("column name" in v for v in vals):
            header_idx = i
            for j, v in enumerate(vals):
                if "column name" in v:
                    col_map["column_name"] = j
                elif "payload final" in v:
                    col_map["payload_final"] = j
                elif "current payload" in v:
                    col_map["current_payload"] = j
                elif "ingestion logic" in v:
                    col_map["ingestion_logic"] = j
                elif "mapping required" in v:
                    col_map["mapping_required"] = j
                elif "remarks" in v:
                    col_map["remarks"] = j
                elif "value present" in v:
                    col_map["value_present"] = j
            break

    if header_idx is None:
        return []

    def _get(row_cells: list, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row_cells):
            return ""
        return _cell(row_cells[idx])

    rules = []
    for row in all_rows[header_idx + 1:]:
        cells = list(row)
        field = _get(cells, "column_name")
        if not field:
            continue
        rules.append({
            "field_name":       field,
            "payload_field":    _get(cells, "payload_final") or _get(cells, "current_payload") or field,
            "ingestion_logic":  _get(cells, "ingestion_logic"),
            "mapping_required": _get(cells, "mapping_required"),
            "remarks":          _get(cells, "remarks"),
            "value_present":    _get(cells, "value_present"),
        })

    return rules


def parse_charge_mapping(excel_bytes: bytes, vendor_ref_id: str) -> list[dict]:
    """
    Open the "Charge Map" sheet and return per-vendor charge code mappings.

    Sheet layout:
      Row 0  : vendor_ref_id column headers  (CMFH, MSGR, CLIM …)
      Row 1  : vendor company names
      Row 2+ : charge data
                 Col 0 = Charge Code (400, DET, 405 …)
                 Col 1 = Master Charge Name
                 Col N = vendor-specific charge name

    Returns:
        [
          {"charge_code": "400", "master_name": "Base Freight",  "vendor_name": "Freight Charge"},
          {"charge_code": "405", "master_name": "Fuel Surcharge","vendor_name": "Fuel Surcharge"},
          ...
        ]
    """
    if not _HAS_OPENPYXL:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if "charge" in name.lower() and "map" in name.lower():
            sheet = wb[name]
            break

    if sheet is None:
        return []

    all_rows = list(sheet.iter_rows())
    if len(all_rows) < 3:
        return []

    # Row 0: vendor IDs — find the column for our vendor
    header = [_cell(c) for c in all_rows[0]]
    vendor_col = next(
        (j for j, v in enumerate(header) if v.strip().upper() == vendor_ref_id.strip().upper()),
        None,
    )

    if vendor_col is None:
        return []

    # Data starts at row 2 (row 1 is company names)
    charges = []
    for row in all_rows[2:]:
        cells = list(row)
        code   = _cell(cells[0]) if cells else ""
        master = _cell(cells[1]) if len(cells) > 1 else ""
        vendor = _cell(cells[vendor_col]) if vendor_col < len(cells) else ""
        if code:
            charges.append({
                "charge_code": code,
                "master_name": master,
                "vendor_name": vendor,
            })

    return charges
